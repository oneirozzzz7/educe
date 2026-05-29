"""
DeepForge 文件处理器
支持：纯文本/代码、PDF、Excel、Word、图片
"""
from __future__ import annotations

import base64
import csv
import io
import mimetypes
import uuid
from dataclasses import dataclass, field
from pathlib import Path

MAX_TEXT_LENGTH = 50000
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".tsx", ".jsx", ".css", ".html", ".htm",
    ".json", ".md", ".yaml", ".yml", ".xml", ".csv", ".sh", ".bash",
    ".sql", ".go", ".java", ".c", ".cpp", ".h", ".hpp", ".rb", ".rs",
    ".swift", ".kt", ".php", ".r", ".m", ".lua", ".pl", ".toml", ".ini",
    ".cfg", ".conf", ".env", ".gitignore", ".dockerfile", ".vue", ".svelte",
    ".scss", ".sass", ".less", ".graphql", ".proto", ".tf", ".log",
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp", ".ico"}

SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | IMAGE_EXTENSIONS | {".pdf", ".xlsx", ".xls", ".docx"}


@dataclass
class FileAttachment:
    id: str
    name: str
    size: int
    mime_type: str
    text_content: str = ""
    is_image: bool = False
    base64_data: str = ""
    error: str = ""

    def to_dict(self) -> dict:
        d = {"id": self.id, "name": self.name, "size": self.size, "mime_type": self.mime_type, "is_image": self.is_image}
        if self.error:
            d["error"] = self.error
        return d


def process_file(path: Path) -> FileAttachment:
    """处理单个文件，提取文本或编码图片"""
    name = path.name
    size = path.stat().st_size
    ext = path.suffix.lower()
    mime = mimetypes.guess_type(name)[0] or "application/octet-stream"
    fid = uuid.uuid4().hex[:10]

    if size > MAX_FILE_SIZE:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error=f"文件过大({size // 1024 // 1024}MB > 10MB)")

    if ext not in SUPPORTED_EXTENSIONS:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error=f"不支持的文件类型: {ext}")

    try:
        if ext in IMAGE_EXTENSIONS:
            return _process_image(path, fid, name, size, mime)
        elif ext == ".pdf":
            return _process_pdf(path, fid, name, size, mime)
        elif ext in (".xlsx", ".xls"):
            return _process_excel(path, fid, name, size, mime)
        elif ext == ".docx":
            return _process_docx(path, fid, name, size, mime)
        else:
            return _process_text(path, fid, name, size, mime)
    except Exception as e:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error=str(e)[:200])


def _process_text(path: Path, fid: str, name: str, size: int, mime: str) -> FileAttachment:
    text = path.read_text(encoding="utf-8", errors="replace")
    if len(text) > MAX_TEXT_LENGTH:
        text = text[:MAX_TEXT_LENGTH] + f"\n\n... [截断: 原文{len(text)}字符, 仅显示前{MAX_TEXT_LENGTH}字符]"
    return FileAttachment(id=fid, name=name, size=size, mime_type=mime, text_content=text)


def _process_pdf(path: Path, fid: str, name: str, size: int, mime: str) -> FileAttachment:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error="需要安装PyPDF2: pip install PyPDF2")

    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- 第{i + 1}页 ---\n{text}")
        if sum(len(p) for p in pages) > MAX_TEXT_LENGTH:
            pages.append(f"\n... [截断: 共{len(reader.pages)}页, 仅提取前{i + 1}页]")
            break

    full_text = "\n\n".join(pages) if pages else "[PDF无法提取文本内容]"
    return FileAttachment(id=fid, name=name, size=size, mime_type=mime, text_content=full_text)


def _process_excel(path: Path, fid: str, name: str, size: int, mime: str) -> FileAttachment:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error="需要安装openpyxl: pip install openpyxl")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    output = io.StringIO()
    writer = csv.writer(output)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.write(f"--- Sheet: {sheet_name} ---\n")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])
            row_count += 1
            if output.tell() > MAX_TEXT_LENGTH:
                output.write(f"\n... [截断: Sheet '{sheet_name}' 共读取{row_count}行]")
                break
        output.write("\n")
        if output.tell() > MAX_TEXT_LENGTH:
            break

    wb.close()
    text = output.getvalue()
    return FileAttachment(id=fid, name=name, size=size, mime_type=mime, text_content=text)


def _process_docx(path: Path, fid: str, name: str, size: int, mime: str) -> FileAttachment:
    try:
        from docx import Document
    except ImportError:
        return FileAttachment(id=fid, name=name, size=size, mime_type=mime, error="需要安装python-docx: pip install python-docx")

    doc = Document(str(path))
    paragraphs = []
    total_len = 0
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append(text)
            total_len += len(text)
            if total_len > MAX_TEXT_LENGTH:
                paragraphs.append(f"\n... [截断: 原文过长, 仅提取前{MAX_TEXT_LENGTH}字符]")
                break

    full_text = "\n\n".join(paragraphs) if paragraphs else "[Word文档无文本内容]"
    return FileAttachment(id=fid, name=name, size=size, mime_type=mime, text_content=full_text)


def _process_image(path: Path, fid: str, name: str, size: int, mime: str) -> FileAttachment:
    data = path.read_bytes()
    b64 = base64.b64encode(data).decode("ascii")
    return FileAttachment(id=fid, name=name, size=size, mime_type=mime, is_image=True, base64_data=b64)


def format_for_prompt(files: list[FileAttachment]) -> str:
    """格式化文件内容为prompt注入段"""
    if not files:
        return ""

    sections = []
    for f in files:
        if f.error:
            sections.append(f"### {f.name} (错误)\n{f.error}")
        elif f.is_image:
            sections.append(f"### {f.name} (图片, {f.size // 1024}KB)\n[图片已传入模型视觉通道]")
        elif f.text_content:
            ext = Path(f.name).suffix.lower().lstrip(".")
            lang = {"py": "python", "js": "javascript", "ts": "typescript", "jsx": "jsx", "tsx": "tsx"}.get(ext, ext)
            if ext in ("csv", "xlsx", "xls"):
                sections.append(f"### {f.name} (表格数据)\n```\n{f.text_content}\n```")
            elif ext == "pdf":
                sections.append(f"### {f.name} (PDF文档)\n{f.text_content}")
            elif ext == "docx":
                sections.append(f"### {f.name} (Word文档)\n{f.text_content}")
            else:
                sections.append(f"### {f.name}\n```{lang}\n{f.text_content}\n```")

    return "\n## 用户上传的文件\n\n" + "\n\n".join(sections)
